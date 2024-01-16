import pandas as pd
import random
from tkinter import filedialog
from console import utils
import os
import time
from openpyxl.styles import Font, Border, Side

def readStudentData():
    global studentNames
    global studentNumbers
    filepath = filedialog.askopenfile(mode='rb', title='请选择学生名单文件', filetype=(("学生名单", "*.xlsx"), ("所有文件", "*.txt")))
    studentInfo = pd.read_excel(io=filepath, sheet_name=0, usecols="A,B").to_dict(orient="list")
    studentNames = studentInfo["姓名"]
    studentNumbers = studentInfo["班级学号"]
    print("成功读取学生名单！")

def queryAboutExamSpots():
    global examSpots
    examSpots = []
    while True:
        try:
            spotCount = int(input("请输入考场总数："))
            if spotCount <= 0:
                print("输入的数值有误，请重新输入！")
                continue
            break
        except ValueError:
            print("输入的数值有误，请重新输入！")

    examSpotsLeft = spotCount
    inputCount = 0
    while examSpotsLeft > 0:
        inputCount += 1
        examSpotInputText = f"请输入第 {inputCount} 个考场的考场名："
        examSpotName = input(examSpotInputText)
        examSpots.append(examSpotName)
        examSpotsLeft -= 1
    print("最终各考场为：")
    print(examSpots)
    print()

def queryAboutStudentInfo():
    global classNo
    global studentInfoPerClass
    global totalClasses
    print("请现在录入考生班级信息：")
    while True:
        try:
            classCount = int(input("请问一共有几个班级的考生："))
            totalClasses = classCount
            break
        except ValueError:
            print("输入有误，请重新输入！")

    classesLeft = classCount
    currentClassCount = 0
    studentInfoPerClass = []
    classNo = []
    while classesLeft > 0:
        currentClassCount += 1
        classNo.append(currentClassCount)
        classCountInputText = f"请输入第 {currentClassCount} 个班级的学生人数："
        while True:
            try:
                currentClassStudentCount = int(input(classCountInputText))
                studentInfoPerClass.append(currentClassStudentCount)
                classesLeft -= 1
                break
            except ValueError:
                print("输入有误，请重新输入!")

    print("最终各班考生人数为：")
    print(studentInfoPerClass)
    print()

def queryAboutHowToDistribute():
    global templateInfo
    print("现在请在打开的窗口内选择考场编排模版文件：")
    try:
        filename = filedialog.askopenfilename(title='请选择一个考场模版文件', filetype=(("模版文件", "*.txt"), ("所有文件", "*.txt")))
        with open(filename, "r") as template:
            templateInfo = template.readline().strip()
            lines, columns = map(int, templateInfo.split(','))
            print(f"成功读取模版！排列方式为 {lines}×{columns}")
    except FileNotFoundError:
        print("文件未找到，请重试！")
    except Exception as e:
        print(f"遇到错误：{e}")

def randomizeTheStudents():
    global randomizedStudentNumbers
    global randomizedStudentNames
    studentCount = 0
    totalStudents = len(studentNames)
    randomizedStudentList = []
    randomizedStudentNumbers = []
    randomizedStudentNames = []

    for student in studentNumbers:
        studentCount += 1
        print(f"正在为第 {studentCount} 个学生排座位！")
        while True:
            studentGlobalPosition = random.randint(1, totalStudents) - 1
            if studentGlobalPosition in randomizedStudentList:
                continue
            else:
                break
        randomizedStudentList.append(studentGlobalPosition)

    for position in randomizedStudentList:
        randomizedStudentNumbers.append(studentNumbers[position])
        randomizedStudentNames.append(studentNames[position])

    print("打乱学生操作完成！\n")


def create_exam_folder():
    if not os.path.exists("考场表"):
        os.makedirs("考场表")


def writeStudentsToExcel(exam_spot_index, student_list, line_coverage, column_coverage):
    df = pd.DataFrame(student_list)
    file_path = f'考场表/ExamSpot_{exam_spot_index + 1}.xlsx'

    # Reorganize students into a seating chart with column_coverage columns and line_coverage rows
    seating_chart = [student_list[i:i + column_coverage * 2] for i in range(0, len(student_list), column_coverage * 2)]

    # Set up Excel writer
    with pd.ExcelWriter(file_path, engine='openpyxl') as excel_writer:
        df.to_excel(excel_writer, index=False, header=['学号', '姓名'])

        # Write exam spot name and total students
        sheet = excel_writer.sheets['Sheet1']
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_coverage + 1)
        sheet.cell(row=1, column=1, value=f'考场 {exam_spot_index + 1}').font = Font(name='宋体', size=12, bold=True)
        sheet.cell(row=1, column=column_coverage + 2, value=f'总人数：{len(student_list)}').font = Font(name='宋体',
                                                                                                       size=12)

        # Write seating chart with specified formatting
        for i, row in enumerate(seating_chart):
            for j, item in enumerate(row):
                sheet.cell(row=i + 2, column=j + 1, value=item).font = Font(name='宋体')
                sheet.cell(row=i + 2, column=j + 1).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Set row height
            sheet.row_dimensions[i + 2].height = 20.4

            # Set column widths for student number and name
            sheet.column_dimensions[pd.get_column_letter(j + 1)].width = 5 if j % 2 == 0 else 11

    print(f"生成考场 {exam_spot_index + 1} 学生列表完成！")


def sortRandomizedStudentsIntoExamSpots():
    lineCoverage, columnCoverage = map(int, templateInfo.split(','))
    create_exam_folder()
    print(f"已选择模板：每个考场排列学生 {columnCoverage} 行， {lineCoverage} 列")
    currentStudentCount = 0
    studentToDistributePerClass = lineCoverage * columnCoverage
    totalTable = []

    for studentName, studentNumber in zip(randomizedStudentNames, randomizedStudentNumbers):
        currentStudentCount += 1
        currentDict = dict(studentNo=studentNumber, student=studentName)
        totalTable.append(currentDict)

        if currentStudentCount % studentToDistributePerClass == 0:
            exam_spot_index = currentStudentCount // studentToDistributePerClass - 1
            print(f"生成考场 {exam_spot_index + 1} 学生列表:")
            print(totalTable)
            writeStudentsToExcel(exam_spot_index, totalTable, lineCoverage, columnCoverage)
            totalTable = []  # 清空列表，为下一考场学生做准备

    if totalTable:  # 处理剩余的学生
        exam_spot_index = (currentStudentCount - 1) // studentToDistributePerClass
        print(f"生成考场 {exam_spot_index + 1} 学生列表:")
        print(totalTable)
        writeStudentsToExcel(exam_spot_index, totalTable, lineCoverage, columnCoverage)

    print("乱序学生分班完成！")

def __main__():
    utils.set_title("自动排考场程序 v1.0")
    readStudentData()
    queryAboutExamSpots()
    queryAboutStudentInfo()
    queryAboutHowToDistribute()
    randomizeTheStudents()
    sortRandomizedStudentsIntoExamSpots()
    try:
        writeStudentsToExcel()
    except:
        pass
    print("\n-----所有操作完成-----\n")
    time.sleep(10)

__main__()
